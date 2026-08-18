from pathlib import Path
from html import escape
from textwrap import wrap
from zipfile import ZIP_DEFLATED, ZipFile
import sqlite3

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "Raw datasets"
DATA_PREP_DIR = ROOT / "Data Preparation"
NUMERIC_DIR = ROOT / "Numeric Analysis"
VIS_DIR = ROOT / "Visualisation"
ANALYSIS_DIR = ROOT / "Python or Excel Data Analysis"
DATABASE_DIR = ROOT / "Database Integration"
REPORT_DIR = ROOT / "Report & Demo"

GROUP_MEMBERS = [
    ("Sibusiso Agent Mathonsi", "202331971"),
    ("Kegoikantse Sebetseba", "202414220"),
    ("Agcobile Qabo", "202469432"),
    ("Lebogang Malatjie", "202404978"),
    ("Tlotlo Naledi", "202422544"),
    ("Tlotlanang Naledi", "202422539"),
]


def group_member_lines() -> list[str]:
    return ["Group Members:"] + [
        f"{index}. {name} {student_number}"
        for index, (name, student_number) in enumerate(GROUP_MEMBERS, start=1)
    ]

IPC_RAW = RAW_DIR / "Integrated Food Security Phase Classification" / "IPC_IPC.csv"
SVFI_RAW = (
    RAW_DIR
    / "Prevalence of severe food insecurity in the population (%)"
    / "WB_WDI_SN_ITK_SVFI_ZS.csv"
)


def ensure_output_dirs() -> None:
    for directory in [DATA_PREP_DIR, NUMERIC_DIR, VIS_DIR, ANALYSIS_DIR, DATABASE_DIR, REPORT_DIR]:
        directory.mkdir(parents=True, exist_ok=True)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy()
    cleaned.columns = (
        cleaned.columns.str.strip()
        .str.lower()
        .str.replace(" ", "_", regex=False)
        .str.replace("/", "_", regex=False)
    )
    return cleaned


def load_raw_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not IPC_RAW.exists():
        raise FileNotFoundError(f"Missing IPC raw dataset: {IPC_RAW}")
    if not SVFI_RAW.exists():
        raise FileNotFoundError(f"Missing severe food insecurity raw dataset: {SVFI_RAW}")

    ipc = normalize_columns(pd.read_csv(IPC_RAW))
    svfi = normalize_columns(pd.read_csv(SVFI_RAW))
    return ipc, svfi


def require_validation(condition: bool, message: str) -> None:
    if not condition:
        raise ValueError(f"Data validation failed: {message}")


def validate_cleaned_data(
    ipc_zaf: pd.DataFrame,
    ipc_phases: pd.DataFrame,
    p3plus: pd.DataFrame,
    svfi_zaf: pd.DataFrame,
) -> pd.DataFrame:
    ipc_missing = ipc_zaf.loc[ipc_zaf["obs_value"].isna()]
    phase5_missing = ipc_missing.loc[
        ipc_missing["phase"].eq("Phase 5 - Famine")
        & ipc_missing["obs_status_label"].eq("Missing value")
    ]

    validation_checks = [
        {
            "control": "Country filter",
            "result": f"{len(ipc_zaf)} IPC rows; {len(svfi_zaf)} SVFI rows",
            "status": "pass",
        },
        {
            "control": "IPC period type",
            "result": ", ".join(sorted(ipc_zaf["time_period"].dropna().unique())),
            "status": "pass",
        },
        {
            "control": "Annual period type",
            "result": f"{int(svfi_zaf['year'].min())}-{int(svfi_zaf['year'].max())}",
            "status": "pass",
        },
        {
            "control": "Missingness and status",
            "result": f"{len(ipc_missing)} missing IPC values; {len(phase5_missing)} Phase 5 missing values",
            "status": "pass",
        },
    ]

    require_validation(len(ipc_zaf) == 12, "expected 12 South Africa IPC rows")
    require_validation(len(svfi_zaf) == 6, "expected 6 South Africa severe-food-insecurity rows")
    require_validation(ipc_zaf["ref_area"].eq("ZAF").all(), "all IPC rows must use REF_AREA = ZAF")
    require_validation(
        svfi_zaf["ref_area"].eq("ZAF").all(),
        "all severe-food-insecurity rows must use REF_AREA = ZAF",
    )
    require_validation(
        ipc_zaf["ref_area_label"].eq("South Africa").all(),
        "all IPC rows must be labelled South Africa",
    )
    require_validation(
        svfi_zaf["ref_area_label"].eq("South Africa").all(),
        "all severe-food-insecurity rows must be labelled South Africa",
    )
    require_validation(ipc_zaf["period"].notna().all(), "all IPC TIME_PERIOD values must parse as YYYY-MM")
    require_validation(
        ipc_zaf["time_period"].eq("2020-10").all(),
        "supplied South Africa IPC rows should all be the October 2020 snapshot",
    )

    svfi_years = svfi_zaf["year"].astype(int).tolist()
    require_validation(
        svfi_years == list(range(2018, 2024)),
        "severe-food-insecurity years must be consecutive from 2018 through 2023",
    )
    require_validation(
        not svfi_zaf.duplicated(["ref_area", "indicator", "time_period"]).any(),
        "severe-food-insecurity business keys must be unique",
    )
    require_validation(
        not ipc_zaf.duplicated(
            ["ref_area", "indicator", "unit_measure", "comp_breakdown_2", "time_period"]
        ).any(),
        "IPC business keys must be unique",
    )

    ipc_people = ipc_zaf.loc[ipc_zaf["metric"].eq("people") & ipc_zaf["obs_value"].notna()]
    ipc_percentages = ipc_zaf.loc[
        ipc_zaf["metric"].eq("percentage") & ipc_zaf["obs_value"].notna()
    ]
    require_validation((ipc_people["obs_value"] >= 0).all(), "IPC people values must be non-negative")
    require_validation(
        ipc_percentages["obs_value"].between(0, 100).all(),
        "IPC percentage values must be between 0 and 100",
    )
    require_validation(
        svfi_zaf["obs_value"].between(0, 100).all(),
        "severe-food-insecurity percentages must be between 0 and 100",
    )
    require_validation(
        len(ipc_missing) == 2
        and len(phase5_missing) == 2
        and set(phase5_missing["metric"]) == {"people", "percentage"},
        "only Phase 5 people and percentage should be missing in the supplied IPC snapshot",
    )
    require_validation(
        set(p3plus["metric"]) == {"people", "percentage"},
        "source-reported IPC Phase 3+ people and percentage rows must both be present",
    )

    reported_phase_people = ipc_phases.loc[
        ipc_phases["metric"].eq("people") & ipc_phases["obs_value"].notna(), "obs_value"
    ].sum()
    reported_phase_percentage = ipc_phases.loc[
        ipc_phases["metric"].eq("percentage") & ipc_phases["obs_value"].notna(), "obs_value"
    ].sum()
    validation_checks.extend(
        [
            {
                "control": "Reported phase totals",
                "result": f"{reported_phase_people:,.0f} people; {reported_phase_percentage:.1f}%",
                "status": "pass",
            },
            {
                "control": "P3+ source rows",
                "result": "Source-reported people and percentage records present",
                "status": "pass",
            },
        ]
    )
    return pd.DataFrame(validation_checks)


def clean_ipc(ipc: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    zaf = ipc.loc[
        (ipc["ref_area"].eq("ZAF")) | (ipc["ref_area_label"].eq("South Africa"))
    ].copy()

    keep_cols = [
        "ref_area",
        "ref_area_label",
        "indicator",
        "indicator_label",
        "unit_measure",
        "unit_measure_label",
        "unit_type",
        "unit_type_label",
        "comp_breakdown_1_label",
        "comp_breakdown_2",
        "comp_breakdown_2_label",
        "time_period",
        "obs_value",
        "obs_status_label",
        "obs_conf_label",
    ]
    zaf = zaf[keep_cols]
    zaf["obs_value"] = pd.to_numeric(zaf["obs_value"], errors="coerce")
    zaf["period"] = pd.to_datetime(zaf["time_period"], format="%Y-%m", errors="coerce")
    zaf["year"] = zaf["period"].dt.year
    zaf["month"] = zaf["period"].dt.month

    zaf["phase"] = (
        zaf["comp_breakdown_2_label"]
        .str.extract(r"(Phase \d - [A-Za-z]+)", expand=False)
        .fillna("Phase 3+")
    )
    zaf["metric"] = np.where(
        zaf["unit_measure_label"].eq("Persons"), "people", "percentage"
    )

    phases = zaf.loc[zaf["indicator"].eq("IPC_IPC_PHASE")].copy()
    phases["phase_order"] = (
        phases["phase"].str.extract(r"Phase (\d)", expand=False).astype(float)
    )
    phases = phases.sort_values(["metric", "phase_order"])

    p3plus = zaf.loc[zaf["indicator"].eq("IPC_IPC_P3PLUS")].copy()
    return zaf, phases, p3plus


def clean_svfi(svfi: pd.DataFrame) -> pd.DataFrame:
    zaf = svfi.loc[
        (svfi["ref_area"].eq("ZAF")) | (svfi["ref_area_label"].eq("South Africa"))
    ].copy()

    keep_cols = [
        "ref_area",
        "ref_area_label",
        "indicator",
        "indicator_label",
        "unit_measure",
        "unit_measure_label",
        "unit_type",
        "unit_type_label",
        "time_period",
        "obs_value",
        "obs_status_label",
        "obs_conf_label",
    ]
    zaf = zaf[keep_cols]
    zaf["year"] = pd.to_numeric(zaf["time_period"], errors="coerce").astype("Int64")
    zaf["obs_value"] = pd.to_numeric(zaf["obs_value"], errors="coerce")
    zaf = zaf.sort_values("year").reset_index(drop=True)
    zaf["year_on_year_change_percentage_points"] = zaf["obs_value"].diff().round(1)
    zaf["rolling_3_year_average"] = zaf["obs_value"].rolling(3, min_periods=3).mean().round(2)
    return zaf


def write_data_preparation_outputs(
    ipc_zaf: pd.DataFrame,
    ipc_phases: pd.DataFrame,
    svfi_zaf: pd.DataFrame,
    validation_results: pd.DataFrame,
) -> None:
    ipc_zaf.to_csv(DATA_PREP_DIR / "cleaned_ipc_south_africa.csv", index=False)
    ipc_phases.to_csv(DATA_PREP_DIR / "cleaned_ipc_phase_distribution_south_africa.csv", index=False)
    svfi_zaf.to_csv(
        DATA_PREP_DIR / "cleaned_severe_food_insecurity_south_africa.csv",
        index=False,
    )

    ipc_people = ipc_zaf.loc[ipc_zaf["metric"].eq("people")]
    ipc_percentages = ipc_zaf.loc[ipc_zaf["metric"].eq("percentage")]
    stats_rows = [
        {
            "dataset": "Integrated Food Security Phase Classification",
            "unit": "people",
            "raw_rows_for_south_africa": len(ipc_people),
            "missing_obs_values": int(ipc_people["obs_value"].isna().sum()),
            "minimum_obs_value": round(float(ipc_people["obs_value"].min()), 2),
            "maximum_obs_value": round(float(ipc_people["obs_value"].max()), 2),
            "mean_obs_value": round(float(ipc_people["obs_value"].mean()), 2),
        },
        {
            "dataset": "Integrated Food Security Phase Classification",
            "unit": "percent",
            "raw_rows_for_south_africa": len(ipc_percentages),
            "missing_obs_values": int(ipc_percentages["obs_value"].isna().sum()),
            "minimum_obs_value": round(float(ipc_percentages["obs_value"].min()), 2),
            "maximum_obs_value": round(float(ipc_percentages["obs_value"].max()), 2),
            "mean_obs_value": round(float(ipc_percentages["obs_value"].mean()), 2),
        },
        {
            "dataset": "Severe food insecurity prevalence",
            "unit": "percent",
            "raw_rows_for_south_africa": len(svfi_zaf),
            "missing_obs_values": int(svfi_zaf["obs_value"].isna().sum()),
            "minimum_obs_value": round(float(svfi_zaf["obs_value"].min()), 2),
            "maximum_obs_value": round(float(svfi_zaf["obs_value"].max()), 2),
            "mean_obs_value": round(float(svfi_zaf["obs_value"].mean()), 2),
        },
    ]
    pd.DataFrame(stats_rows).to_csv(
        DATA_PREP_DIR / "descriptive_statistics.csv", index=False
    )
    validation_results.to_csv(DATA_PREP_DIR / "data_validation_results.csv", index=False)

    notes = [
        "DATA QUALITY NOTES",
        "",
        "Scope:",
        "- Both raw datasets were filtered to South Africa using REF_AREA = ZAF or REF_AREA_LABEL = South Africa.",
        "- Raw files were not edited. Cleaned outputs were saved in the Data Preparation folder.",
        "",
        "Integrated Food Security Phase Classification:",
        f"- South Africa rows retained: {len(ipc_zaf)}.",
        f"- Missing OBS_VALUE entries: {int(ipc_zaf['obs_value'].isna().sum())}.",
        "- Missing IPC Phase 5 values were kept as missing in the cleaned file and summary tables.",
        "- Phase 5 is labelled Not reported in stakeholder outputs; it is not converted to 0.",
        "- The IPC dataset provides a snapshot for 2020-10, including counts and percentages by phase.",
        "",
        "Severe food insecurity prevalence:",
        f"- South Africa rows retained: {len(svfi_zaf)}.",
        f"- Missing OBS_VALUE entries: {int(svfi_zaf['obs_value'].isna().sum())}.",
        "- Values are percentages and cover the available years in the raw dataset.",
        "- The 3-year rolling average is published from 2020 onward; 2018 and 2019 remain blank because fewer than three observations are available.",
    ]
    (DATA_PREP_DIR / "data_quality_notes.txt").write_text("\n".join(notes), encoding="utf-8")


def write_numeric_outputs(
    ipc_phases: pd.DataFrame, p3plus: pd.DataFrame, svfi_zaf: pd.DataFrame
) -> tuple[pd.DataFrame, dict[str, float]]:
    people = ipc_phases.loc[ipc_phases["metric"].eq("people")].copy()
    percentages = ipc_phases.loc[ipc_phases["metric"].eq("percentage")].copy()

    reported_phase_people = people["obs_value"].sum()
    reported_phase_percentage = percentages["obs_value"].sum()
    p3plus_people = float(p3plus.loc[p3plus["metric"].eq("people"), "obs_value"].iloc[0])
    p3plus_percent = float(p3plus.loc[p3plus["metric"].eq("percentage"), "obs_value"].iloc[0])
    p3plus_values = {"people": p3plus_people, "percentage": p3plus_percent}
    phase_row_crisis_or_worse_people = people.loc[
        people["phase"].isin(["Phase 3 - Crisis", "Phase 4 - Emergency", "Phase 5 - Famine"]),
        "obs_value",
    ].sum()
    missing_phase_values = int(ipc_phases["obs_value"].isna().sum())

    svfi_values = svfi_zaf["obs_value"].to_numpy(dtype=float)
    svfi_years = svfi_zaf["year"].to_numpy(dtype=int)
    svfi_changes = np.diff(svfi_values)

    summary = pd.DataFrame(
        [
            {
                "metric": "IPC reported phase-row population, 2020-10",
                "value": round(reported_phase_people, 2),
                "unit": "people",
            },
            {
                "metric": "IPC reported phase-row share, 2020-10",
                "value": round(reported_phase_percentage, 2),
                "unit": "percent",
            },
            {
                "metric": "IPC source-reported people in Phase 3 or above, 2020-10",
                "value": round(p3plus_people, 2),
                "unit": "people",
            },
            {
                "metric": "IPC source-reported percentage in Phase 3 or above, 2020-10",
                "value": round(p3plus_percent, 2),
                "unit": "percent",
            },
            {
                "metric": "Reported Phase 3-4 people from phase rows, 2020-10",
                "value": round(float(phase_row_crisis_or_worse_people), 2),
                "unit": "people",
            },
            {
                "metric": "IPC missing phase observations retained",
                "value": missing_phase_values,
                "unit": "records",
            },
            {
                "metric": "Severe food insecurity prevalence average",
                "value": round(float(np.mean(svfi_values)), 2),
                "unit": "percent",
            },
            {
                "metric": "Severe food insecurity prevalence minimum",
                "value": round(float(np.min(svfi_values)), 2),
                "unit": "percent",
            },
            {
                "metric": "Severe food insecurity prevalence maximum",
                "value": round(float(np.max(svfi_values)), 2),
                "unit": "percent",
            },
            {
                "metric": "Severe food insecurity total change",
                "value": round(float(svfi_values[-1] - svfi_values[0]), 2),
                "unit": "percentage points",
            },
        ]
    )
    summary.to_csv(NUMERIC_DIR / "numerical_summary.csv", index=False)

    yoy = pd.DataFrame(
        {
            "from_year": svfi_years[:-1],
            "to_year": svfi_years[1:],
            "change_percentage_points": np.round(svfi_changes, 2),
        }
    )
    yoy.to_csv(NUMERIC_DIR / "severe_food_insecurity_year_on_year_change.csv", index=False)

    phase_summary = people[["phase", "phase_order", "obs_value", "obs_status_label"]].merge(
        percentages[["phase", "phase_order", "obs_value", "obs_status_label"]],
        on=["phase", "phase_order"],
        suffixes=("_people", "_percentage"),
    )
    phase_summary.rename(
        columns={
            "obs_value_people": "people",
            "obs_status_label_people": "people_status",
            "obs_value_percentage": "percentage",
            "obs_status_label_percentage": "percentage_status",
        },
        inplace=True,
    )
    phase_summary["people"] = phase_summary["people"].round(0)
    phase_summary["percentage"] = phase_summary["percentage"].round(1)
    phase_summary["reporting_treatment"] = np.where(
        phase_summary[["people", "percentage"]].isna().any(axis=1),
        "Not reported in source",
        "Reported",
    )
    phase_summary["people_display"] = phase_summary["people"].map(
        lambda value: "Not reported" if pd.isna(value) else f"{value:,.0f}"
    )
    phase_summary["percentage_display"] = phase_summary["percentage"].map(
        lambda value: "Not reported" if pd.isna(value) else f"{value:.1f}%"
    )
    phase_summary = phase_summary.sort_values("phase_order").reset_index(drop=True)
    phase_summary.to_csv(NUMERIC_DIR / "ipc_phase_summary.csv", index=False)

    findings = [
        "NUMERICAL FINDINGS",
        "",
        f"- The severe food insecurity prevalence series runs from {svfi_years[0]} to {svfi_years[-1]}.",
        f"- Prevalence increased from {svfi_values[0]:.1f}% in {svfi_years[0]} to {svfi_values[-1]:.1f}% in {svfi_years[-1]}, a change of {svfi_values[-1] - svfi_values[0]:.1f} percentage points.",
        f"- The highest severe food insecurity prevalence in the available data is {np.max(svfi_values):.1f}%.",
        f"- The average severe food insecurity prevalence is {np.mean(svfi_values):.2f}%.",
        f"- In the IPC 2020-10 snapshot, {p3plus_people:,.0f} people were in Phase 3 or above, equal to {p3plus_percent:.1f}% of the classified population.",
        f"- Reported Phase 3 and Phase 4 rows sum to {phase_row_crisis_or_worse_people:,.0f} people; Phase 5 is not reported in the supplied phase rows.",
        "- Use the source-reported Phase 3+ row as the authoritative crisis-or-worse KPI.",
    ]
    (NUMERIC_DIR / "numerical_findings.txt").write_text("\n".join(findings), encoding="utf-8")

    return phase_summary, p3plus_values


def write_analysis_summary(
    svfi_zaf: pd.DataFrame, phase_summary: pd.DataFrame, p3plus_values: dict[str, float]
) -> None:
    latest_svfi = svfi_zaf.iloc[-1]
    analysis_summary = pd.DataFrame(
        [
            {
                "section": "Severe food insecurity trend",
                "finding": f"South Africa's severe food insecurity prevalence reached {latest_svfi['obs_value']:.1f}% in {int(latest_svfi['year'])}, the latest supplied source observation.",
            },
            {
                "section": "IPC phase distribution",
                "finding": f"The IPC 2020-10 snapshot source-reports {p3plus_values['people']:,.0f} people, or {p3plus_values['percentage']:.1f}%, in Phase 3 or above.",
            },
            {
                "section": "IPC Phase 5 limitation",
                "finding": "Phase 5 is not reported in the supplied IPC phase rows and should not be displayed as zero.",
            },
            {
                "section": "Project interpretation",
                "finding": "The datasets can be used together to explain both longer-term severe food insecurity prevalence and a detailed phase snapshot.",
            },
        ]
    )
    analysis_summary.to_csv(ANALYSIS_DIR / "analysis_summary.csv", index=False)
    phase_summary.to_csv(ANALYSIS_DIR / "transformed_ipc_phase_table.csv", index=False)
    svfi_zaf.to_csv(ANALYSIS_DIR / "transformed_svfi_trend_table.csv", index=False)


def write_excel_workbook(
    svfi_zaf: pd.DataFrame, phase_summary: pd.DataFrame, p3plus_values: dict[str, float]
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    members_sheet = workbook.create_sheet("Group Members")
    members_sheet.append(["No.", "Name", "Student number"])
    for index, (name, student_number) in enumerate(GROUP_MEMBERS, start=1):
        members_sheet.append([index, name, student_number])

    summary = workbook.create_sheet("Summary")
    summary.append(["Section", "Finding"])
    summary.append([
        "Severe food insecurity trend",
        f"Prevalence increased from {svfi_zaf.iloc[0]['obs_value']:.1f}% in {int(svfi_zaf.iloc[0]['year'])} to {svfi_zaf.iloc[-1]['obs_value']:.1f}% in {int(svfi_zaf.iloc[-1]['year'])}.",
    ])
    summary.append([
        "IPC phase snapshot",
        f"The 2020-10 IPC snapshot source-reports {p3plus_values['percentage']:.1f}% of the classified population in Phase 3 or above.",
    ])
    summary.append([
        "IPC Phase 5 limitation",
        "Phase 5 people and percentage are not reported in the supplied phase rows; do not display them as zero.",
    ])
    summary.append([
        "Interpretation",
        "The trend dataset shows worsening severe food insecurity, while IPC gives a detailed severity distribution for one snapshot.",
    ])

    svfi_sheet = workbook.create_sheet("SVFI Trend")
    svfi_sheet.append(["Year", "Prevalence (%)", "Year-on-year change", "3-year rolling average"])
    for row in svfi_zaf[["year", "obs_value", "year_on_year_change_percentage_points", "rolling_3_year_average"]].itertuples(index=False):
        svfi_sheet.append(list(row))

    line_chart = LineChart()
    line_chart.title = "Severe food insecurity prevalence"
    line_chart.y_axis.title = "Percent"
    line_chart.x_axis.title = "Year"
    line_chart.add_data(Reference(svfi_sheet, min_col=2, min_row=1, max_row=svfi_sheet.max_row), titles_from_data=True)
    line_chart.set_categories(Reference(svfi_sheet, min_col=1, min_row=2, max_row=svfi_sheet.max_row))
    svfi_sheet.add_chart(line_chart, "F2")

    phase_sheet = workbook.create_sheet("IPC Phase Snapshot")
    phase_sheet.append(
        ["Phase", "People", "Percentage", "People display", "Percentage display", "Reporting treatment"]
    )
    for row in phase_summary[
        [
            "phase",
            "people",
            "percentage",
            "people_display",
            "percentage_display",
            "reporting_treatment",
        ]
    ].itertuples(index=False):
        phase_sheet.append(list(row))

    bar_chart = BarChart()
    bar_chart.title = "IPC phase distribution by percentage"
    bar_chart.y_axis.title = "Percent"
    bar_chart.x_axis.title = "Phase"
    phase_chart_rows = int(phase_summary["percentage"].notna().sum()) + 1
    bar_chart.add_data(Reference(phase_sheet, min_col=3, min_row=1, max_row=phase_chart_rows), titles_from_data=True)
    bar_chart.set_categories(Reference(phase_sheet, min_col=1, min_row=2, max_row=phase_chart_rows))
    phase_sheet.add_chart(bar_chart, "E2")

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 55)

    workbook.save(ANALYSIS_DIR / "food_security_analysis_workbook.xlsx")


def write_database_outputs(
    ipc_zaf: pd.DataFrame, ipc_phases: pd.DataFrame, svfi_zaf: pd.DataFrame, phase_summary: pd.DataFrame
) -> None:
    database_path = DATABASE_DIR / "food_security.db"
    if database_path.exists():
        database_path.unlink()

    with sqlite3.connect(database_path) as connection:
        ipc_zaf.to_sql("ipc_cleaned", connection, index=False, if_exists="replace")
        ipc_phases.to_sql("ipc_phase_distribution", connection, index=False, if_exists="replace")
        svfi_zaf.to_sql("severe_food_insecurity", connection, index=False, if_exists="replace")
        phase_summary.to_sql("ipc_phase_summary", connection, index=False, if_exists="replace")

        connection.execute("CREATE INDEX idx_svfi_year ON severe_food_insecurity(year)")
        connection.execute("CREATE INDEX idx_ipc_phase_metric ON ipc_phase_distribution(phase, metric)")

        p3plus_result = pd.read_sql_query(
            """
            SELECT
                MAX(CASE WHEN metric = 'people' THEN obs_value END) AS source_reported_p3plus_people,
                MAX(CASE WHEN metric = 'percentage' THEN obs_value END) AS source_reported_p3plus_percentage
            FROM ipc_cleaned
            WHERE indicator = 'IPC_IPC_P3PLUS';
            """,
            connection,
        )
        svfi_result = pd.read_sql_query(
            """
            SELECT year, obs_value AS severe_food_insecurity_percent,
                   year_on_year_change_percentage_points,
                   rolling_3_year_average
            FROM severe_food_insecurity
            ORDER BY year;
            """,
            connection,
        )
        phase_result = pd.read_sql_query(
            """
            SELECT phase, people, percentage, reporting_treatment
            FROM ipc_phase_summary
            ORDER BY phase_order;
            """,
            connection,
        )
        update_demo = pd.DataFrame(
            [
                {
                    "operation": "UPDATE demonstration",
                    "sql_pattern": "UPDATE severe_food_insecurity SET obs_status_label = 'Reviewed for project' WHERE year = 2023;",
                    "safety_note": "Uses a WHERE clause and should be run only in a copied/demo table for assessment evidence.",
                },
                {
                    "operation": "DELETE demonstration",
                    "sql_pattern": "DELETE FROM severe_food_insecurity WHERE year IS NULL;",
                    "safety_note": "Uses a WHERE clause that targets invalid rows only; no valid rows in this project match it.",
                },
            ]
        )

    p3plus_result.to_csv(DATABASE_DIR / "query_result_crisis_or_worse.csv", index=False)
    svfi_result.to_csv(DATABASE_DIR / "query_result_svfi_trend.csv", index=False)
    phase_result.to_csv(DATABASE_DIR / "query_result_ipc_phase_rank.csv", index=False)
    update_demo.to_csv(DATABASE_DIR / "safe_update_delete_evidence.csv", index=False)

    schema = """
DROP TABLE IF EXISTS ipc_cleaned;
DROP TABLE IF EXISTS ipc_phase_distribution;
DROP TABLE IF EXISTS severe_food_insecurity;
DROP TABLE IF EXISTS ipc_phase_summary;

-- Tables are populated by the Python pipeline from cleaned CSV/DataFrame outputs.
-- The schema below documents the table purpose and key analytical fields.
CREATE TABLE ipc_cleaned (
    ref_area TEXT,
    ref_area_label TEXT,
    indicator TEXT,
    indicator_label TEXT,
    unit_measure TEXT,
    unit_measure_label TEXT,
    unit_type TEXT,
    unit_type_label TEXT,
    comp_breakdown_1_label TEXT,
    comp_breakdown_2 TEXT,
    comp_breakdown_2_label TEXT,
    time_period TEXT,
    obs_value REAL,
    obs_status_label TEXT,
    obs_conf_label TEXT,
    period TEXT,
    year INTEGER,
    month INTEGER,
    phase TEXT,
    metric TEXT
);

CREATE TABLE ipc_phase_distribution (
    ref_area TEXT,
    ref_area_label TEXT,
    indicator TEXT,
    indicator_label TEXT,
    unit_measure TEXT,
    unit_measure_label TEXT,
    unit_type TEXT,
    unit_type_label TEXT,
    comp_breakdown_1_label TEXT,
    comp_breakdown_2 TEXT,
    comp_breakdown_2_label TEXT,
    time_period TEXT,
    obs_value REAL,
    obs_status_label TEXT,
    obs_conf_label TEXT,
    period TEXT,
    year INTEGER,
    month INTEGER,
    phase TEXT,
    metric TEXT,
    phase_order REAL
);

CREATE TABLE severe_food_insecurity (
    ref_area TEXT,
    ref_area_label TEXT,
    indicator TEXT,
    indicator_label TEXT,
    unit_measure TEXT,
    unit_measure_label TEXT,
    unit_type TEXT,
    unit_type_label TEXT,
    time_period INTEGER,
    obs_value REAL,
    obs_status_label TEXT,
    obs_conf_label TEXT,
    year INTEGER,
    year_on_year_change_percentage_points REAL,
    rolling_3_year_average REAL
);

CREATE TABLE ipc_phase_summary (
    phase TEXT,
    phase_order REAL,
    people REAL,
    people_status TEXT,
    percentage REAL,
    percentage_status TEXT,
    reporting_treatment TEXT,
    people_display TEXT,
    percentage_display TEXT
);
""".strip()
    (DATABASE_DIR / "schema.sql").write_text(schema, encoding="utf-8")

    queries = """
-- 1. Severe food insecurity trend by year.
SELECT year, obs_value AS severe_food_insecurity_percent, year_on_year_change_percentage_points
FROM severe_food_insecurity
ORDER BY year;

-- 2. IPC phase distribution in phase order, with missing Phase 5 preserved.
SELECT phase, people, percentage, reporting_treatment
FROM ipc_phase_summary
ORDER BY phase_order;

-- 3. Source-reported Phase 3+ KPI for Crisis or worse.
SELECT
    MAX(CASE WHEN metric = 'people' THEN obs_value END) AS source_reported_p3plus_people,
    MAX(CASE WHEN metric = 'percentage' THEN obs_value END) AS source_reported_p3plus_percentage
FROM ipc_cleaned
WHERE indicator = 'IPC_IPC_P3PLUS';

-- 4. Phase-row reconciliation. SUM ignores the missing Phase 5 value; this is not a Phase 5 zero estimate.
SELECT
    SUM(people) AS reported_phase_3_4_people,
    SUM(percentage) AS reported_phase_3_4_percentage,
    SUM(CASE WHEN people IS NULL OR percentage IS NULL THEN 1 ELSE 0 END) AS missing_phase_rows
FROM ipc_phase_summary
WHERE phase IN ('Phase 3 - Crisis', 'Phase 4 - Emergency', 'Phase 5 - Famine');

-- 5. Safe UPDATE example for a copied/demo table.
CREATE TABLE IF NOT EXISTS severe_food_insecurity_demo AS
SELECT * FROM severe_food_insecurity;
UPDATE severe_food_insecurity_demo
SET obs_status_label = 'Reviewed for project'
WHERE year = 2023;

-- 6. Safe DELETE example targeting invalid records only.
DELETE FROM severe_food_insecurity_demo
WHERE year IS NULL;
""".strip()
    (DATABASE_DIR / "queries.sql").write_text(queries, encoding="utf-8")

    notes = [
        "DATABASE INTEGRATION EVIDENCE",
        "",
        "- food_security.db is a SQLite database generated from the cleaned project outputs.",
        "- schema.sql documents the analytical tables.",
        "- queries.sql contains SELECT queries plus safe UPDATE and DELETE examples using WHERE clauses.",
        "- query_result_*.csv files are exported query outputs for report evidence.",
        "- safe_update_delete_evidence.csv explains the update/delete safety controls.",
    ]
    (DATABASE_DIR / "database_evidence_notes.txt").write_text("\n".join(notes), encoding="utf-8")


def report_sections(
    svfi_zaf: pd.DataFrame, phase_summary: pd.DataFrame, p3plus_values: dict[str, float]
) -> list[tuple[str, list[str]]]:
    start_year = int(svfi_zaf.iloc[0]["year"])
    end_year = int(svfi_zaf.iloc[-1]["year"])
    start_value = float(svfi_zaf.iloc[0]["obs_value"])
    end_value = float(svfi_zaf.iloc[-1]["obs_value"])
    reported_phase_3_4_people = phase_summary.loc[
        phase_summary["phase"].isin(["Phase 3 - Crisis", "Phase 4 - Emergency", "Phase 5 - Famine"]),
        "people",
    ].sum()
    p3plus = p3plus_values["people"]
    p3plus_pct = p3plus_values["percentage"]

    return [
        (
            "Executive Summary",
            [
                "This project analyses South African food-security conditions using two World Bank Data360 datasets: the Integrated Food Security Phase Classification (IPC) and the World Development Indicators severe food insecurity prevalence series.",
                f"The main finding is that severe food insecurity prevalence rose from {start_value:.1f}% in {start_year} to {end_value:.1f}% in {end_year}. The IPC 2020-10 snapshot also shows {p3plus:,.0f} people, or {p3plus_pct:.1f}% of the classified population, in Phase 3 or above.",
                "Together, the datasets describe both a multi-year trend and a detailed severity distribution. The trend suggests worsening pressure over time, while the IPC snapshot identifies a substantial Crisis-or-worse population at one point in time.",
            ],
        ),
        (
            "Data Sources and Scope",
            [
                "The IPC dataset contains monthly classification records by country, indicator, unit, phase, and time period. For this project it was filtered to South Africa records only.",
                "The severe food insecurity prevalence dataset contains annual World Development Indicators observations. For South Africa, the available records cover 2018 to 2023.",
                "The raw CSV files were preserved unchanged in the Raw datasets folder. All cleaning, transformations, and outputs were written to separate project folders for auditability.",
            ],
        ),
        (
            "Data Preparation",
            [
                "Column names were normalised into lowercase snake_case fields. South Africa was selected using REF_AREA = ZAF or REF_AREA_LABEL = South Africa.",
                "Numeric observation values were converted with explicit coercion, time periods were parsed into year/month fields, and IPC phase labels were extracted into readable phase names.",
                "The IPC Phase 5 records contain missing values for both people and percentage. These missing values remain visible in the cleaned file, summary tables, workbook, database, and charts as Not reported.",
            ],
        ),
        (
            "Numerical Analysis",
            [
                f"The severe food insecurity average across the available period is {svfi_zaf['obs_value'].mean():.2f}%. The minimum is {svfi_zaf['obs_value'].min():.1f}% and the maximum is {svfi_zaf['obs_value'].max():.1f}%.",
                f"The total change over the observed period is {end_value - start_value:.1f} percentage points. Year-on-year changes are exported in the Numeric Analysis folder.",
                f"The IPC distribution shows Phase 1 and Phase 2 as the largest reported phase groups. Reported Phase 3 and Phase 4 rows sum to {reported_phase_3_4_people:,.0f} people, while the source-reported Phase 3+ KPI remains {p3plus:,.0f} people.",
            ],
        ),
        (
            "Database Integration",
            [
                "A SQLite database named food_security.db was generated from the cleaned datasets. It includes tables for cleaned IPC records, IPC phase distribution, severe food insecurity trend data, and summarised IPC phase results.",
                "The database folder includes schema.sql, queries.sql, exported SELECT query results, and safe UPDATE/DELETE examples. The update and delete examples use WHERE clauses and are intended for a copied demo table.",
                "This demonstrates that the cleaned data can be stored, queried, modified safely, and exported for reporting or visualisation.",
            ],
        ),
        (
            "Visualisation",
            [
                "The visualisation folder contains a severe food insecurity trend line chart, IPC phase distribution charts by people and percentage, and a combined dashboard.",
                "The charts use clear titles, axis labels, units, value labels, and consistent colours. They are designed for direct use in the final report and class demonstration.",
                "The trend line makes the rise from 2018 to 2023 visible, while the IPC bars show the population scale of each classification phase.",
            ],
        ),
        (
            "Conclusion and Recommendations",
            [
                "South Africa's severe food insecurity prevalence increased across the observed WDI series, ending at its highest supplied value in 2023.",
                "Policy and programme discussion should use the source-reported Phase 3+ IPC result as a planning trigger, not as a stand-alone allocation rule.",
                "Recommended next steps are to refresh the source data, add household-level or provincial evidence where available, and document comparability before making operational claims.",
            ],
        ),
    ]


def write_report_markdown(sections: list[tuple[str, list[str]]]) -> None:
    lines = [
        "# NDTA631 Food Security Data Analysis and Visualisation Report",
        "",
        "Topic: South African food security and severe food insecurity",
        "",
        "Prepared for: NDTA631 Data Analysis and Visualisation Group Assignment 2026",
        "",
        *group_member_lines(),
        "",
    ]
    for heading, paragraphs in sections:
        lines.append(f"## {heading}")
        lines.extend([""])
        for paragraph in paragraphs:
            lines.extend([paragraph, ""])
    lines.extend(
        [
            "## Report Figures",
            "",
            "- Visualisation/severe_food_insecurity_trend.png",
            "- Visualisation/ipc_phase_distribution_people.png",
            "- Visualisation/ipc_phase_distribution_percentage.png",
            "- Visualisation/food_security_overview_dashboard.png",
            "",
            "## Repository Evidence",
            "",
            "- Data Preparation: cleaned CSV files, descriptive statistics, and data quality notes.",
            "- Numeric Analysis: NumPy-driven summary tables and numerical findings.",
            "- Python or Excel Data Analysis: reusable Python pipeline and Excel workbook.",
            "- Database Integration: SQLite database, schema, queries, and exported query results.",
            "- Visualisation: final PNG charts and chart explanations.",
        ]
    )
    (REPORT_DIR / "NDTA631_Group_Report.md").write_text("\n".join(lines), encoding="utf-8")


def docx_paragraph(text: str, style: str | None = None) -> str:
    style_xml = f'<w:pStyle w:val="{style}"/>' if style else ""
    return (
        "<w:p><w:pPr>"
        f"{style_xml}"
        "</w:pPr><w:r><w:t xml:space=\"preserve\">"
        f"{escape(text)}"
        "</w:t></w:r></w:p>"
    )


def docx_image(rel_id: str, width_emu: int = 5486400, height_emu: int = 3200400) -> str:
    return f"""
<w:p>
  <w:r>
    <w:drawing>
      <wp:inline distT="0" distB="0" distL="0" distR="0" xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing">
        <wp:extent cx="{width_emu}" cy="{height_emu}"/>
        <wp:docPr id="1" name="Project chart"/>
        <a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
          <a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">
            <pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">
              <pic:nvPicPr><pic:cNvPr id="0" name="chart.png"/><pic:cNvPicPr/></pic:nvPicPr>
              <pic:blipFill><a:blip r:embed="{rel_id}"/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>
              <pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{width_emu}" cy="{height_emu}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>
            </pic:pic>
          </a:graphicData>
        </a:graphic>
      </wp:inline>
    </w:drawing>
  </w:r>
</w:p>
""".strip()


def write_docx_report(sections: list[tuple[str, list[str]]]) -> None:
    chart_files = [
        VIS_DIR / "severe_food_insecurity_trend.png",
        VIS_DIR / "ipc_phase_distribution_people.png",
        VIS_DIR / "ipc_phase_distribution_percentage.png",
        VIS_DIR / "food_security_overview_dashboard.png",
    ]

    body = [
        docx_paragraph("NDTA631 Food Security Data Analysis and Visualisation Report", "Title"),
        docx_paragraph("South African food security and severe food insecurity", "Subtitle"),
        docx_paragraph("Prepared for NDTA631 Group Assignment 2026"),
        *(docx_paragraph(line) for line in group_member_lines()),
    ]
    for heading, paragraphs in sections:
        body.append(docx_paragraph(heading, "Heading1"))
        body.extend(docx_paragraph(paragraph) for paragraph in paragraphs)
    body.append(docx_paragraph("Report Figures", "Heading1"))
    for index, image_path in enumerate(chart_files, start=1):
        if image_path.exists():
            body.append(docx_paragraph(image_path.name, "Heading2"))
            body.append(docx_image(f"rId{index + 1}"))

    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas"
xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"
xmlns:o="urn:schemas-microsoft-com:office:office"
xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math"
xmlns:v="urn:schemas-microsoft-com:vml"
xmlns:wp14="http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing"
xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
xmlns:w10="urn:schemas-microsoft-com:office:word"
xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml"
xmlns:wpg="http://schemas.microsoft.com/office/word/2010/wordprocessingGroup"
xmlns:wpi="http://schemas.microsoft.com/office/word/2010/wordprocessingInk"
xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml"
xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape"
mc:Ignorable="w14 wp14"><w:body>{''.join(body)}<w:sectPr><w:pgSz w:w="12240" w:h="15840"/><w:pgMar w:top="720" w:right="720" w:bottom="720" w:left="720"/></w:sectPr></w:body></w:document>"""

    relationships = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">',
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>',
    ]
    for index, image_path in enumerate(chart_files, start=1):
        if image_path.exists():
            relationships.append(
                f'<Relationship Id="rId{index + 1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/{image_path.name}"/>'
            )
    relationships.append("</Relationships>")

    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:style w:type="paragraph" w:default="1" w:styleId="Normal"><w:name w:val="Normal"/><w:rPr><w:sz w:val="22"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/><w:rPr><w:b/><w:sz w:val="34"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Subtitle"><w:name w:val="Subtitle"/><w:rPr><w:i/><w:sz w:val="24"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/><w:rPr><w:b/><w:sz w:val="28"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Heading2"><w:name w:val="heading 2"/><w:rPr><w:b/><w:sz w:val="24"/></w:rPr></w:style>
</w:styles>"""

    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Default Extension="png" ContentType="image/png"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
</Types>"""

    package_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""

    with ZipFile(REPORT_DIR / "NDTA631_Group_Report.docx", "w", ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml", content_types)
        docx.writestr("_rels/.rels", package_rels)
        docx.writestr("word/document.xml", document_xml)
        docx.writestr("word/styles.xml", styles_xml)
        docx.writestr("word/_rels/document.xml.rels", "\n".join(relationships))
        for image_path in chart_files:
            if image_path.exists():
                docx.write(image_path, f"word/media/{image_path.name}")


def add_pdf_text_page(pdf: PdfPages, title: str, paragraphs: list[str]) -> None:
    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor("white")
    y = 0.94
    fig.text(0.08, y, title, fontsize=16, weight="bold", va="top")
    y -= 0.05
    for paragraph in paragraphs:
        for line in wrap(paragraph, width=88):
            fig.text(0.08, y, line, fontsize=10.5, va="top")
            y -= 0.023
        y -= 0.018
        if y < 0.08:
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)
            fig = plt.figure(figsize=(8.27, 11.69))
            fig.patch.set_facecolor("white")
            y = 0.94
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def write_pdf_report(sections: list[tuple[str, list[str]]]) -> None:
    chart_files = [
        VIS_DIR / "severe_food_insecurity_trend.png",
        VIS_DIR / "ipc_phase_distribution_people.png",
        VIS_DIR / "ipc_phase_distribution_percentage.png",
        VIS_DIR / "food_security_overview_dashboard.png",
    ]
    with PdfPages(REPORT_DIR / "NDTA631_Group_Report.pdf") as pdf:
        add_pdf_text_page(
            pdf,
            "NDTA631 Food Security Data Analysis and Visualisation Report",
            [
                "South African food security and severe food insecurity.",
                "Prepared for NDTA631 Group Assignment 2026.",
                *group_member_lines(),
            ],
        )
        for heading, paragraphs in sections:
            add_pdf_text_page(pdf, heading, paragraphs)
        for image_path in chart_files:
            if image_path.exists():
                image = plt.imread(image_path)
                fig, ax = plt.subplots(figsize=(11.69, 8.27))
                ax.imshow(image)
                ax.axis("off")
                fig.suptitle(image_path.name, fontsize=14, weight="bold")
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)


def write_demo_materials(sections: list[tuple[str, list[str]]]) -> None:
    script = [
        "NDTA631 PROJECT DEMO SCRIPT",
        "",
        *group_member_lines(),
        "",
        "1. Open README.md and explain the project question, folder structure, and data source.",
        "2. Show the raw datasets folder and state that the original CSV files were preserved unchanged.",
        "3. Run: python \"Python or Excel Data Analysis\\food_security_pipeline.py\"",
        "4. Open Data Preparation outputs and explain South Africa filtering, missing values, and cleaned files.",
        "5. Open Numeric Analysis outputs and explain the severe food insecurity trend and IPC Phase 3+ result.",
        "6. Open Database Integration/food_security.db or the exported query CSV files, then explain schema.sql and queries.sql.",
        "7. Open the Visualisation folder and discuss the trend chart, IPC phase charts, and dashboard.",
        "8. Open Report & Demo/NDTA631_Group_Report.pdf and use it as the final story for the audience.",
        "",
        "Key talking point:",
        sections[0][1][1],
    ]
    (REPORT_DIR / "demo_script.txt").write_text("\n".join(script), encoding="utf-8")

    checklist = [
        "FINAL SUBMISSION CHECKLIST",
        "",
        "[x] Group member names and student numbers added to the DOCX/PDF.",
        "[x] Raw datasets preserved unchanged.",
        "[x] Cleaned South Africa datasets created.",
        "[x] Missing values and cleaning decisions documented.",
        "[x] NumPy/Pandas numerical analysis completed.",
        "[x] Excel workbook created with analysis tables and charts.",
        "[x] SQLite database, schema, queries, and query results created.",
        "[x] Visualisations exported as PNG files.",
        "[x] Report draft generated in Markdown, DOCX, and PDF.",
        "[x] Demo script prepared.",
    ]
    (REPORT_DIR / "final_submission_checklist.txt").write_text("\n".join(checklist), encoding="utf-8")


def write_report_outputs(
    svfi_zaf: pd.DataFrame, phase_summary: pd.DataFrame, p3plus_values: dict[str, float]
) -> None:
    sections = report_sections(svfi_zaf, phase_summary, p3plus_values)
    write_report_markdown(sections)
    write_docx_report(sections)
    write_pdf_report(sections)
    write_demo_materials(sections)


def save_visualisations(svfi_zaf: pd.DataFrame, phase_summary: pd.DataFrame) -> None:
    sns.set_theme(style="whitegrid", palette="Set2")
    plt.rcParams["figure.dpi"] = 140

    fig, ax = plt.subplots(figsize=(9, 5))
    sns.lineplot(data=svfi_zaf, x="year", y="obs_value", marker="o", linewidth=2.5, ax=ax)
    ax.set_title("South Africa: Severe Food Insecurity Prevalence", fontsize=14, weight="bold")
    ax.set_xlabel("Year")
    ax.set_ylabel("Population with severe food insecurity (%)")
    ax.set_xticks(svfi_zaf["year"].astype(int).tolist())
    ax.set_ylim(0, max(10, svfi_zaf["obs_value"].max() + 1))
    for _, row in svfi_zaf.iterrows():
        ax.text(row["year"], row["obs_value"] + 0.12, f"{row['obs_value']:.1f}%", ha="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(VIS_DIR / "severe_food_insecurity_trend.png")
    plt.close(fig)

    plot_people = phase_summary.copy()
    plot_people["people_millions"] = plot_people["people"] / 1_000_000
    plot_people["people_millions_plot"] = plot_people["people_millions"].fillna(0)
    plot_people["percentage_plot"] = plot_people["percentage"].fillna(0)
    phase_colors = np.where(
        plot_people["reporting_treatment"].eq("Reported"), "#4C78A8", "#BDBDBD"
    )
    y_positions = np.arange(len(plot_people))

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(y_positions, plot_people["people_millions_plot"], color=phase_colors)
    ax.set_title("South Africa IPC Phase Distribution by People, 2020-10", fontsize=14, weight="bold")
    ax.set_xlabel("People (millions)")
    ax.set_ylabel("IPC phase")
    ax.set_yticks(y_positions, plot_people["phase"])
    ax.invert_yaxis()
    ax.set_xlim(0, max(plot_people["people_millions_plot"].max() * 1.18, 1))
    label_offset = max(plot_people["people_millions_plot"].max() * 0.015, 0.1)
    for y, row in enumerate(plot_people.itertuples(index=False)):
        label = "Not reported" if pd.isna(row.people) else f"{row.people_millions:.1f}M"
        ax.text(row.people_millions_plot + label_offset, y, label, va="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(VIS_DIR / "ipc_phase_distribution_people.png")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(y_positions, plot_people["percentage_plot"], color=phase_colors)
    ax.set_title("South Africa IPC Phase Distribution by Percentage, 2020-10", fontsize=14, weight="bold")
    ax.set_xlabel("IPC phase")
    ax.set_ylabel("Classified population (%)")
    ax.set_xticks(y_positions, plot_people["phase"], rotation=25, ha="right")
    ax.set_ylim(0, max(plot_people["percentage_plot"].max() * 1.18, 1))
    pct_label_offset = max(plot_people["percentage_plot"].max() * 0.015, 0.3)
    for x, row in enumerate(plot_people.itertuples(index=False)):
        label = "Not reported" if pd.isna(row.percentage) else f"{row.percentage:.0f}%"
        ax.text(x, row.percentage_plot + pct_label_offset, label, ha="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(VIS_DIR / "ipc_phase_distribution_percentage.png")
    plt.close(fig)

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    sns.lineplot(data=svfi_zaf, x="year", y="obs_value", marker="o", ax=axes[0])
    axes[0].set_title("Severe Food Insecurity Trend")
    axes[0].set_xlabel("Year")
    axes[0].set_ylabel("% of population")
    axes[1].barh(y_positions, plot_people["people_millions_plot"], color=phase_colors)
    axes[1].set_title("IPC Phase Snapshot, 2020-10")
    axes[1].set_xlabel("People (millions)")
    axes[1].set_ylabel("")
    axes[1].set_yticks(y_positions, plot_people["phase"])
    axes[1].invert_yaxis()
    axes[1].set_xlim(0, max(plot_people["people_millions_plot"].max() * 1.18, 1))
    for y, row in enumerate(plot_people.itertuples(index=False)):
        label = "Not reported" if pd.isna(row.people) else f"{row.people_millions:.1f}M"
        axes[1].text(row.people_millions_plot + label_offset, y, label, va="center", fontsize=8)
    fig.suptitle("South Africa Food Security Overview", fontsize=16, weight="bold")
    fig.tight_layout()
    fig.savefig(VIS_DIR / "food_security_overview_dashboard.png")
    plt.close(fig)

    explanations = [
        "CHART EXPLANATIONS",
        "",
        "severe_food_insecurity_trend.png:",
        "- Shows the annual percentage of South Africa's population experiencing severe food insecurity in the available WDI data.",
        "- The line rises from 2018 to 2023, indicating a worsening prevalence over the period.",
        "",
        "ipc_phase_distribution_people.png:",
        "- Shows the number of people in each IPC phase in the 2020-10 South Africa snapshot.",
        "- Phase 5 is labelled Not reported because the supplied source value is missing.",
        "",
        "ipc_phase_distribution_percentage.png:",
        "- Shows the same IPC phase distribution as percentages of the classified population.",
        "- The chart keeps Phase 5 as Not reported and uses the source-reported Phase 3+ result for the crisis-or-worse KPI.",
        "",
        "food_security_overview_dashboard.png:",
        "- Combines the trend view and IPC snapshot for use in the report or presentation.",
    ]
    (VIS_DIR / "chart_explanations.txt").write_text("\n".join(explanations), encoding="utf-8")


def main() -> None:
    ensure_output_dirs()
    ipc_raw, svfi_raw = load_raw_data()
    ipc_zaf, ipc_phases, p3plus = clean_ipc(ipc_raw)
    svfi_zaf = clean_svfi(svfi_raw)
    validation_results = validate_cleaned_data(ipc_zaf, ipc_phases, p3plus, svfi_zaf)
    write_data_preparation_outputs(ipc_zaf, ipc_phases, svfi_zaf, validation_results)
    phase_summary, p3plus_values = write_numeric_outputs(ipc_phases, p3plus, svfi_zaf)
    write_analysis_summary(svfi_zaf, phase_summary, p3plus_values)
    write_excel_workbook(svfi_zaf, phase_summary, p3plus_values)
    write_database_outputs(ipc_zaf, ipc_phases, svfi_zaf, phase_summary)
    save_visualisations(svfi_zaf, phase_summary)
    write_report_outputs(svfi_zaf, phase_summary, p3plus_values)
    print("Food security analysis pipeline completed successfully.")


if __name__ == "__main__":
    main()
